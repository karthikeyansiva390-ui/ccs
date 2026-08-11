
from __future__ import annotations

import io
import math
import re
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import numpy_financial as npf
import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# CCS Screening Engine
#
# Important design principle:
# The field workbook is NOT compared directly against the reference workbook.
# Field values first pass through a normalization/canonicalization layer.
#
# Numerical:
#   raw field text -> missing-value detection -> number + unit extraction
#   -> unit conversion to reference unit -> canonical numeric value
#   -> hard-cutoff test -> SAW-range lookup
#
# Qualitative:
#   raw field text -> missing-value detection -> Unicode/spacing/punctuation
#   normalization -> exact normalized reference match -> conservative fuzzy
#   match against HARD CUTOFF + ALL SAW TEXT VALUES -> canonical reference text
#   -> hard-cutoff test -> SAW text classification
#
# This prevents differences such as:
#   "SUITABLE", "Suitable", "suitable", "Suitable."
# from being treated as different values.
# ---------------------------------------------------------------------------

CATEGORIES = [
    "Technical",
    "Environmental",
    "Regulatory",
    "Long Term Operation",
    "Risk",
]

MISSING_MARKERS = {
    "",
    "-",
    "--",
    "---",
    "—",
    "–",
    "n/a",
    "na",
    "n.a",
    "n.a.",
    "not available",
    "not applicable",
    "not provided",
    "unknown",
    "none",
    "null",
    "nil",
    "blank",
    "missing",
    "not specified",
    "not reported",
}

def norm_text(x: Any) -> str:
    """Unicode-safe, case-insensitive normalization for text matching."""
    if x is None:
        return ""
    s = str(x).replace("\u00a0", " ").replace("\u2013", "-").replace("\u2014", "-")
    s = s.strip().casefold()
    s = re.sub(r"[\r\n\t]+", " ", s)
    s = re.sub(r"[“”‘’]", "'", s)
    s = re.sub(r"[^0-9a-zA-ZÀ-ÖØ-öø-ÿ%/().,+\- ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def is_missing(x: Any) -> bool:
    if x is None:
        return True
    if isinstance(x, float) and math.isnan(x):
        return True
    s = norm_text(x)
    return s in MISSING_MARKERS


def canonical_missing(x: Any) -> Optional[str]:
    return None if is_missing(x) else str(x).strip()


def split_choices(x: Any) -> List[str]:
    if is_missing(x):
        return []
    return [v.strip() for v in re.split(r"[|;]", str(x)) if not is_missing(v)]


def clean_number_str(s: str) -> str:
    # Remove currency symbols and thousands separators, but retain decimal/sign.
    s = s.replace(",", "").replace("$", "").replace("€", "").replace("£", "")
    return s


def parse_number_strict(x: Any) -> Optional[float]:
    """
    Extract a numeric value only if the string is fundamentally numeric.
    This intentionally rejects arbitrary text such as 'approximately high',
    'unknown 20', or '20 mD and good quality'.
    """
    if is_missing(x):
        return None
    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)

    s = str(x).strip()
    s2 = clean_number_str(s)
    # Numeric value with optional scientific notation and an optional unit suffix.
    pattern = r"^[\s]*[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?[\s%a-zA-Zµμ/.*^0-9_-]*$"
    if not re.match(pattern, s2):
        return None
    m = re.match(r"^[\s]*([+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?)", s2)
    if not m:
        return None
    try:
        return float(m.group(1))
    except Exception:
        return None


def extract_unit_text(x: Any) -> str:
    if is_missing(x) or isinstance(x, (int, float, np.integer, np.floating)):
        return ""
    s = clean_number_str(str(x).strip())
    m = re.match(
        r"^[\s]*[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?\s*(.*)$",
        s,
    )
    return (m.group(1) if m else "").strip().casefold()


def canonical_unit(unit: Any) -> str:
    u = norm_text(unit)
    if not u:
        return ""
    return UNIT_ALIASES.get(u, u)


def to_quantity(value: float, unit: str):
    unit = canonical_unit(unit)
    if not unit:
        return value * ureg.dimensionless
    try:
        return value * ureg(unit)
    except Exception:
        return None


def convert_numeric_to_reference(
    field_value: Any,
    field_unit: Any,
    reference_unit: Any,
) -> Tuple[Optional[float], str]:
    """
    Simple framework rule: do not perform automatic unit conversion.
    The reference datasheet and manually entered field value are compared
    directly using the reference unit/type shown in the UI.
    """
    value = parse_number_strict(field_value)
    if value is None:
        return None, "not a valid numeric value"

    return value, f"direct comparison using reference unit/type '{reference_unit}'"


def parse_operator_cutoff(value: Any) -> Tuple[Optional[str], Optional[float]]:
    if is_missing(value):
        return None, None
    if isinstance(value, (int, float, np.integer, np.floating)):
        return ">", float(value)
    s = clean_number_str(str(value))
    m = re.match(
        r"^\s*(>=|<=|>|<|=)\s*([+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?)",
        s,
    )
    if m:
        return m.group(1), float(m.group(2))
    n = parse_number_strict(s)
    return (">", n) if n is not None else (None, None)


def numeric_cutoff_pass(field_value: float, cutoff: float, operator: str) -> bool:
    if operator == ">":
        return field_value > cutoff
    if operator == ">=":
        return field_value >= cutoff
    if operator == "<":
        return field_value < cutoff
    if operator == "<=":
        return field_value <= cutoff
    if operator == "=":
        return math.isclose(field_value, cutoff, rel_tol=1e-9, abs_tol=1e-9)
    return field_value > cutoff


def parse_saw_score_numeric(
    field_value: float,
    ranges: List[Tuple[Any, Any, int]],
) -> Optional[int]:
    """
    Inclusive SAW range matching. Boundary behavior:
      [a,b] is inclusive.
    If adjacent ranges share a boundary, the first range wins.
    """
    for start, end, score in ranges:
        a = parse_number_strict(start)
        b = parse_number_strict(end)
        if a is None and b is None:
            continue
        if a is not None and b is None and math.isclose(field_value, a, rel_tol=1e-9, abs_tol=1e-9):
            return score
        if a is None and b is not None and math.isclose(field_value, b, rel_tol=1e-9, abs_tol=1e-9):
            return score
        if a is not None and b is not None:
            lo, hi = min(a, b), max(a, b)
            if lo <= field_value <= hi:
                return score
    return None


def normalize_phrase_for_matching(x: Any) -> str:
    """Simple case-insensitive reference-text normalization; no fuzzy matching."""
    s = norm_text(x)
    s = re.sub(r"[\\.,;:!?()\\[\\]{}]", " ", s)
    s = re.sub(r"[-_/]", " ", s)
    s = re.sub(r"\\s+", " ", s).strip()
    return s


def parse_saw_score_qualitative(
    field_value: Any,
    ranges: List[Tuple[Any, Any, int]],
) -> Dict[str, Any]:
    """
    Match the manually selected qualitative text against the reference SAW
    text/sentence values. No fuzzy/semantic matching is used.
    Capitalization and harmless punctuation/spacing differences are ignored.
    """
    raw_norm = normalize_phrase_for_matching(field_value)

    for start_value, end_value, score in ranges:
        for value in (start_value, end_value):
            if is_missing(value):
                continue
            for choice in split_choices(value) or [str(value)]:
                if normalize_phrase_for_matching(choice) == raw_norm:
                    return {
                        "score": score,
                        "canonical": choice,
                        "similarity": 100.0,
                        "method": "normalized-exact-reference-match",
                    }

    return {
        "score": None,
        "canonical": None,
        "similarity": 0.0,
        "method": "not-found-in-five-saw-text-values",
    }


def _row_values(ws, row: int, start_col: int = 1, end_col: int = 14) -> List[Any]:
    return [ws.cell(row, c).value for c in range(start_col, end_col + 1)]


def find_reference_header(ws) -> Optional[int]:
    for r in range(1, ws.max_row + 1):
        row = [norm_text(v) for v in _row_values(ws, r)]
        if "parameter" in row and any("hard cut off" in v for v in row):
            return r
    return None


def weight_fraction(x: Any) -> Optional[float]:
    """
    Convert an AHP weight from the reference datasheet to a decimal fraction.

    Examples:
      0.4241      -> 0.4241
      42.41       -> 0.4241
      "42.41%"    -> 0.4241
      "0.4241"    -> 0.4241
    """
    if is_missing(x):
        return None

    if isinstance(x, (int, float, np.integer, np.floating)):
        v = float(x)
    else:
        s = clean_number_str(str(x).strip())
        has_percent = "%" in s
        s = s.replace("%", "").strip()
        try:
            v = float(s)
        except (TypeError, ValueError):
            return None
        if has_percent or abs(v) > 1:
            v /= 100.0

    return v if math.isfinite(v) and v >= 0 else None


def parse_reference_xlsx(file_bytes: bytes) -> Dict[str, Any]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheets = wb.sheetnames

    if not all(cat in sheets for cat in CATEGORIES):
        missing = [cat for cat in CATEGORIES if cat not in sheets]
        raise ValueError("Reference XLSX is missing: " + ", ".join(missing))

    overall_weights: Dict[str, float] = {}
    if "Overall" in sheets:
        ws = wb["Overall"]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if norm_text(ws.cell(r, 1).value) == "parameter" and "ahp" in norm_text(ws.cell(r, 2).value):
                header_row = r
                break
        if header_row:
            for r in range(header_row + 1, ws.max_row + 1):
                name = ws.cell(r, 1).value
                if not name:
                    continue
                match = next((c for c in CATEGORIES if norm_text(c) == norm_text(name)), None)
                if match:
                    w = weight_fraction(ws.cell(r, 2).value)
                    if w is not None:
                        overall_weights[match] = w

    if not overall_weights:
        raise ValueError("No overall AHP weights were found in the Overall sheet.")
    total = sum(overall_weights.values())
    if total <= 0:
        raise ValueError("Overall AHP weights must be positive.")
    overall_weights = {k: v / total for k, v in overall_weights.items()}

    reference = {
        "overall_weights": overall_weights,
        "categories": {},
        "source_type": "xlsx",
    }

    for cat in CATEGORIES:
        ws = wb[cat]
        header_row = find_reference_header(ws)
        if header_row is None:
            raise ValueError(f"Could not identify reference header in '{cat}'.")

        rows = []
        for r in range(header_row + 2, ws.max_row + 1):
            parameter = ws.cell(r, 1).value
            if is_missing(parameter):
                continue
            dtype = ws.cell(r, 2).value
            cutoff = ws.cell(r, 3).value
            ahp = ws.cell(r, 4).value

            ranges = [
                (ws.cell(r, 5).value, ws.cell(r, 6).value, 1),
                (ws.cell(r, 7).value, ws.cell(r, 8).value, 2),
                (ws.cell(r, 9).value, ws.cell(r, 10).value, 3),
                (ws.cell(r, 11).value, ws.cell(r, 12).value, 4),
                (ws.cell(r, 13).value, ws.cell(r, 14).value, 5),
            ]

            rows.append(
                {
                    "parameter": str(parameter).strip(),
                    "data_type": str(dtype).strip() if dtype is not None else "",
                    "hard_cutoff": cutoff,
                    "ahp_weight": weight_fraction(ahp),
                    "saw_ranges": ranges,
                }
            )

        if not rows:
            raise ValueError(f"No parameters found in '{cat}'.")

        sub_sum = sum(r["ahp_weight"] or 0 for r in rows)
        if sub_sum <= 0:
            raise ValueError(f"Sub-parameter AHP weights for '{cat}' are invalid.")
        for r in rows:
            r["ahp_weight_normalized"] = (r["ahp_weight"] or 0) / sub_sum

        reference["categories"][cat] = rows

    return reference


def parse_reference_csv(file_bytes: bytes) -> Dict[str, Any]:
    df = pd.read_csv(io.BytesIO(file_bytes))
    df.columns = [str(c).strip() for c in df.columns]
    lower = {norm_text(c): c for c in df.columns}

    group_col = lower.get("parameter group") or lower.get("category") or lower.get("group")
    param_col = lower.get("parameter")
    ahp_col = next((c for k, c in lower.items() if "ahp" in k and "weight" in k), None)
    cutoff_col = next((c for k, c in lower.items() if "hard cut" in k), None)

    if not group_col or not param_col or not ahp_col or not cutoff_col:
        raise ValueError(
            "Reference CSV needs Parameter Group, Parameter, Hard Cut Off Values and AHP Weightage columns."
        )

    def score_col(score: int, side: str):
        patterns = [
            f"saw score - {score} {side}",
            f"saw score {score} {side}",
            f"saw {score} {side}",
        ]
        for p in patterns:
            if p in lower:
                return lower[p]
        return None

    overall = {}
    for _, row in df.iterrows():
        group = norm_text(row[group_col])
        param = norm_text(row[param_col])
        if group == "overall" and param in {norm_text(c) for c in CATEGORIES}:
            w = weight_fraction(row[ahp_col])
            if w is not None:
                canonical = next(c for c in CATEGORIES if norm_text(c) == param)
                overall[canonical] = w

    if not overall:
        raise ValueError("Reference CSV has no Overall AHP weight rows.")
    total = sum(overall.values())
    overall = {k: v / total for k, v in overall.items()}

    reference = {"overall_weights": overall, "categories": {}, "source_type": "csv"}

    for cat in CATEGORIES:
        cdf = df[df[group_col].astype(str).map(norm_text) == norm_text(cat)]
        if cdf.empty:
            raise ValueError(f"Reference CSV has no rows for {cat}.")

        rows = []
        for _, row in cdf.iterrows():
            ranges = []
            for score in range(1, 6):
                sc = score_col(score, "start")
                ec = score_col(score, "end")
                ranges.append(
                    (
                        row[sc] if sc else None,
                        row[ec] if ec else None,
                        score,
                    )
                )
            rows.append(
                {
                    "parameter": str(row[param_col]).strip(),
                    "data_type": str(row.get(lower.get("data type", ""), "")).strip(),
                    "hard_cutoff": row[cutoff_col],
                    "ahp_weight": weight_fraction(row[ahp_col]),
                    "saw_ranges": ranges,
                }
            )
        sub_sum = sum(r["ahp_weight"] or 0 for r in rows)
        if sub_sum <= 0:
            raise ValueError(f"Sub-parameter AHP weights for {cat} are invalid.")
        for r in rows:
            r["ahp_weight_normalized"] = (r["ahp_weight"] or 0) / sub_sum
        reference["categories"][cat] = rows

    return reference


def parse_reference(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    return parse_reference_csv(file_bytes) if filename.lower().endswith(".csv") else parse_reference_xlsx(file_bytes)


def parse_field_xlsx(file_bytes: bytes) -> Dict[str, Any]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    values = {c: {} for c in CATEGORIES}
    units = {c: {} for c in CATEGORIES}
    economics = {}
    economics_units = {}

    for cat in CATEGORIES:
        if cat not in wb.sheetnames:
            continue
        ws = wb[cat]
        header = None
        for r in range(1, ws.max_row + 1):
            if norm_text(ws.cell(r, 1).value) == "parameter" and norm_text(ws.cell(r, 2).value) == "value":
                header = r
                break
        if header is None:
            continue
        for r in range(header + 1, ws.max_row + 1):
            p = ws.cell(r, 1).value
            if is_missing(p):
                continue
            key = norm_text(p)
            values[cat][key] = ws.cell(r, 2).value
            units[cat][key] = ws.cell(r, 3).value

    if "Economics" in wb.sheetnames:
        ws = wb["Economics"]
        header = None
        for r in range(1, ws.max_row + 1):
            if norm_text(ws.cell(r, 1).value) == "parameter" and norm_text(ws.cell(r, 2).value) == "value":
                header = r
                break
        if header:
            for r in range(header + 1, ws.max_row + 1):
                p = ws.cell(r, 1).value
                if is_missing(p):
                    continue
                key = norm_text(p)
                economics[key] = ws.cell(r, 2).value
                economics_units[key] = ws.cell(r, 3).value

    return {
        "values": values,
        "units": units,
        "economics": economics,
        "economics_units": economics_units,
        "source_type": "xlsx",
    }


def parse_field_csv(file_bytes: bytes, filename: str = "") -> Dict[str, Any]:
    df = pd.read_csv(io.BytesIO(file_bytes))
    df.columns = [str(c).strip() for c in df.columns]
    lower = {norm_text(c): c for c in df.columns}

    param_col = lower.get("parameter")
    value_col = lower.get("value")
    group_col = lower.get("parameter group") or lower.get("category") or lower.get("group")
    unit_col = lower.get("unit/type") or lower.get("unit") or lower.get("data type")

    if not param_col or not value_col or not group_col:
        raise ValueError("Field CSV requires Parameter, Value and Parameter Group columns.")

    values = {c: {} for c in CATEGORIES}
    units = {c: {} for c in CATEGORIES}
    economics, economics_units = {}, {}

    for _, row in df.iterrows():
        if is_missing(row[param_col]):
            continue
        group = str(row[group_col]).strip()
        p = norm_text(row[param_col])
        val = row[value_col]
        unit = row[unit_col] if unit_col else ""

        if norm_text(group) == "economics":
            economics[p] = val
            economics_units[p] = unit
        else:
            cat = next((c for c in CATEGORIES if norm_text(c) == norm_text(group)), None)
            if cat:
                values[cat][p] = val
                units[cat][p] = unit

    return {
        "values": values,
        "units": units,
        "economics": economics,
        "economics_units": economics_units,
        "source_type": "csv",
    }


def parse_field(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    return parse_field_csv(file_bytes, filename) if filename.lower().endswith(".csv") else parse_field_xlsx(file_bytes)


def find_field_value(field_data: Dict[str, Any], category: str, parameter: str) -> Tuple[Any, Any]:
    values = field_data["values"].get(category, {})
    units = field_data["units"].get(category, {})
    target = normalize_phrase_for_matching(parameter)

    for key, value in values.items():
        if normalize_phrase_for_matching(key) == target:
            return value, units.get(key)

    return None, None


def evaluate_parameter(
    field_value: Any,
    field_unit: Any,
    spec: Dict[str, Any],
) -> Dict[str, Any]:
    """
    Phase-1 rule requested by the user:

    1. Read the manually entered value.
    2. For numerical values, FIRST search the five SAW ranges.
       For qualitative values, FIRST search the five SAW text values.
    3. If the value is not represented by any SAW score, check the hard
       cut-off as a fallback.
    4. If the hard cut-off is also not satisfied, eliminate the field.
    5. If hard cut-off passes but no SAW reference match exists, PASS without a SAW score. Only a real SAW match receives a score and AHP weighting.

    No fuzzy matching and no automatic unit conversion are performed.
    """
    if is_missing(field_value):
        return {
            "pass": False,
            "reason": f"{spec['parameter']}: input is missing (blank, N/A, Nil or '-').",
            "normalized_value": None,
            "normalization_note": "missing",
            "saw_score": None,
            "saw_match": None,
            "saw_similarity": None,
            "weighted_score": None,
        }

    dtype = norm_text(spec.get("data_type"))
    is_qualitative = dtype == "qualitative"

    # ---------------- Numerical / measurable ----------------
    if not is_qualitative:
        value = parse_number_strict(field_value)
        if value is None:
            return {
                "pass": False,
                "reason": f"{spec['parameter']}: invalid numerical value.",
                "normalized_value": None,
                "normalization_note": "manual input is not numeric",
                "saw_score": None,
                "saw_match": None,
                "saw_similarity": None,
                "weighted_score": None,
            }

        # FIRST: search the five SAW ranges.
        saw = parse_saw_score_numeric(value, spec["saw_ranges"])
        if saw is not None:
            weighted = saw * float(spec["ahp_weight_normalized"])
            return {
                "pass": True,
                "reason": "",
                "normalized_value": value,
                "normalization_note": "direct numerical comparison; no unit conversion",
                "saw_score": saw,
                "saw_match": f"SAW Score {saw} reference range",
                "saw_similarity": 100.0,
                "weighted_score": weighted,
            }

        # SECOND: SAW did not contain the value, so test the hard cut-off.
        cutoff = spec["hard_cutoff"]
        op, threshold = parse_operator_cutoff(cutoff)
        if threshold is None:
            return {
                "pass": False,
                "reason": f"{spec['parameter']}: value was outside all five SAW ranges and the reference hard cut-off is not a valid numerical criterion.",
                "normalized_value": value,
                "normalization_note": "SAW-first; hard-cutoff fallback unavailable",
                "saw_score": None,
                "saw_match": None,
                "saw_similarity": None,
                "weighted_score": None,
            }

        if not numeric_cutoff_pass(value, threshold, op or ">"):
            return {
                "pass": False,
                "reason": (
                    f"{spec['parameter']}: value {value:g} {spec.get('data_type','')} "
                    f"did not fall within any SAW range and did not satisfy hard cut-off "
                    f"'{op or '>'} {threshold:g}'."
                ),
                "normalized_value": value,
                "normalization_note": "SAW-first; hard-cutoff fallback failed",
                "saw_score": None,
                "saw_match": None,
                "saw_similarity": None,
                "weighted_score": None,
            }

        # It passed the fallback gate, but no SAW score exists. Never invent a
        # score. Keep it out of the final ranked score rather than assigning 0.
        return {
            "pass": True,
            "reason": "",
            "normalized_value": value,
            "normalization_note": "SAW-first; hard-cutoff fallback passed; no SAW score assigned",
            "saw_score": None,
            "saw_match": None,
            "saw_similarity": None,
            "weighted_score": None,
        }

    # ---------------- Qualitative ----------------
    # FIRST: exact normalized match against every SAW text/sentence.
    saw = parse_saw_score_qualitative(field_value, spec["saw_ranges"])
    if saw["score"] is not None:
        weighted = saw["score"] * float(spec["ahp_weight_normalized"])
        return {
            "pass": True,
            "reason": "",
            "normalized_value": saw["canonical"],
            "normalization_note": "direct reference-text match; capitalization/punctuation ignored",
            "saw_score": saw["score"],
            "saw_match": saw["canonical"],
            "saw_similarity": 100.0,
            "weighted_score": weighted,
        }

    # SECOND: no SAW text match -> check hard-cutoff text.
    cutoff_values = split_choices(spec["hard_cutoff"]) or [str(spec["hard_cutoff"])]
    field_norm = normalize_phrase_for_matching(field_value)
    cutoff_match = next(
        (v for v in cutoff_values if normalize_phrase_for_matching(v) == field_norm),
        None,
    )

    if cutoff_match is None:
        return {
            "pass": False,
            "reason": (
                f"{spec['parameter']}: input text did not match any of the five "
                f"SAW reference texts and did not satisfy the hard-cutoff text."
            ),
            "normalized_value": None,
            "normalization_note": "SAW-first; hard-cutoff fallback failed",
            "saw_score": None,
            "saw_match": None,
            "saw_similarity": 0.0,
            "weighted_score": None,
        }

    # Hard cutoff is satisfied, but because no SAW category was matched, do not
    # fabricate a score.
    return {
        "pass": True,
        "reason": "",
        "normalized_value": cutoff_match,
        "normalization_note": "SAW-first; hard-cutoff fallback passed; no SAW score assigned",
        "saw_score": None,
        "saw_match": None,
        "saw_similarity": 100.0,
        "weighted_score": None,
    }


def screen_one_field(field_name: str, field_data: Dict[str, Any], reference: Dict[str, Any]) -> Dict[str, Any]:
    category_scores = {}
    parameter_details = {}
    reasons = []

    for cat in CATEGORIES:
        details = []
        weighted_sum = 0.0

        for spec in reference["categories"][cat]:
            value, unit = find_field_value(field_data, cat, spec["parameter"])
            result = evaluate_parameter(value, unit, spec)

            details.append(
                {
                    "Parameter": spec["parameter"],
                    "Raw Field Value": value,
                    "Field Unit": unit,
                    "Normalized / Canonical Value": result["normalized_value"],
                    "Normalization": result["normalization_note"],
                    "Hard Cut-Off": spec["hard_cutoff"],
                    "Sub-AHP Weight": spec["ahp_weight_normalized"],
                    "SAW Score": result["saw_score"],
                    "Matched SAW Reference": result["saw_match"],
                    "SAW Match Similarity (%)": result["saw_similarity"],
                    "Weighted SAW": result["weighted_score"],
                    "Status": "PASS" if result["pass"] else "FAIL",
                }
            )

            if not result["pass"]:
                reasons.append(f"{cat} — {result['reason']}")
            else:
                weighted_sum += result["weighted_score"]

        category_scores[cat] = weighted_sum
        parameter_details[cat] = details

    passed = not reasons
    overall_score = None
    if passed:
        overall_score = sum(
            category_scores[c] * reference["overall_weights"].get(c, 0)
            for c in CATEGORIES
        )

    actual_capex = parse_number_strict(field_data["economics"].get(norm_text("Actual CAPEX")))
    actual_opex = parse_number_strict(field_data["economics"].get(norm_text("Actual OPEX")))

    return {
        "Field": field_name,
        "Passed": passed,
        "Overall SAW Score": overall_score,
        "Technical": category_scores.get("Technical"),
        "Environmental": category_scores.get("Environmental"),
        "Regulatory": category_scores.get("Regulatory"),
        "Long Term Operation": category_scores.get("Long Term Operation"),
        "Risk": category_scores.get("Risk"),
        "Actual CAPEX": actual_capex,
        "Actual OPEX": actual_opex,
        "Reasons": reasons,
        "Parameter Details": parameter_details,
        "Field Data": field_data,
    }


def screen_fields(reference: Dict[str, Any], field_files: List[Tuple[str, bytes]]):
    passed, failed, details = [], [], {}

    for filename, file_bytes in field_files:
        field_name = re.sub(r"\.(xlsx|xls|csv)$", "", filename, flags=re.I)
        data = parse_field(file_bytes, filename)
        result = screen_one_field(field_name, data, reference)
        details[field_name] = result
        (passed if result["Passed"] else failed).append(result)

    passed.sort(key=lambda x: x["Overall SAW Score"], reverse=True)

    ranked_rows = []
    for rank, r in enumerate(passed, start=1):
        ranked_rows.append(
            {
                "Rank": rank,
                "Field": r["Field"],
                "Overall SAW Score": round(r["Overall SAW Score"], 4),
                "Technical": round(r["Technical"], 4),
                "Environmental": round(r["Environmental"], 4),
                "Regulatory": round(r["Regulatory"], 4),
                "Long Term Operation": round(r["Long Term Operation"], 4),
                "Risk": round(r["Risk"], 4),
            }
        )

    failed_rows = [
        {"Field": r["Field"], "Reason": "\n".join(f"• {x}" for x in r["Reasons"])}
        for r in failed
    ]

    return pd.DataFrame(ranked_rows), pd.DataFrame(failed_rows), details


# ---------------------------------------------------------------------------
# Phase 2 functions retained from the original framework implementation.
# ---------------------------------------------------------------------------

def economic_gate(selected_fields, details, expected_capex, expected_opex):
    passed, failed = [], []
    for field in selected_fields:
        r = details[field]
        reasons = []
        if r.get("Actual CAPEX") is None:
            reasons.append("Actual CAPEX is missing or invalid.")
        elif not expected_capex > r["Actual CAPEX"]:
            reasons.append(
                f"Expected CAPEX ({expected_capex:,.2f}) is not greater than actual CAPEX ({r['Actual CAPEX']:,.2f})."
            )
        if r.get("Actual OPEX") is None:
            reasons.append("Actual OPEX is missing or invalid.")
        elif not expected_opex > r["Actual OPEX"]:
            reasons.append(
                f"Expected OPEX ({expected_opex:,.2f}) is not greater than actual OPEX ({r['Actual OPEX']:,.2f})."
            )
        row = {
            "Field": field,
            "Actual CAPEX": r.get("Actual CAPEX"),
            "Actual OPEX": r.get("Actual OPEX"),
            "Expected CAPEX": expected_capex,
            "Expected OPEX": expected_opex,
            "Gate Status": "PASS" if not reasons else "ELIMINATED",
            "Reason": "\n".join(f"• {x}" for x in reasons),
        }
        (passed if not reasons else failed).append(row)
    return pd.DataFrame(passed), pd.DataFrame(failed)


def project_cashflows(
    actual_capex,
    actual_opex,
    carbon_credit,
    government_subsidy,
    tax_incentive,
    storage_fee,
    discount_rate_pct,
    inflation_rate_pct,
    project_lifetime,
    injection_rate_mtpa,
    carbon_price,
    inflate_revenues=True,
):
    if actual_capex is None or actual_opex is None:
        raise ValueError("Actual CAPEX and OPEX are required.")
    if project_lifetime < 1 or injection_rate_mtpa <= 0:
        raise ValueError("Project lifetime must be >=1 year and injection rate >0.")

    discount = discount_rate_pct / 100.0
    inflation = inflation_rate_pct / 100.0
    tonnes = injection_rate_mtpa * 1_000_000.0
    unit_revenue = carbon_credit + government_subsidy + tax_incentive + storage_fee + carbon_price

    cashflows = [-actual_capex]
    for year in range(1, int(project_lifetime) + 1):
        factor = (1 + inflation) ** (year - 1)
        revenue_factor = factor if inflate_revenues else 1.0
        revenue = tonnes * unit_revenue * revenue_factor
        opex = actual_opex * factor
        cashflows.append(revenue - opex)

    npv = float(sum(cf / ((1 + discount) ** i) for i, cf in enumerate(cashflows)))

    try:
        irr = float(npf.irr(cashflows))
        if not np.isfinite(irr):
            irr = None
    except Exception:
        irr = None

    cumulative = cashflows[0]
    payback = None
    for year in range(1, len(cashflows)):
        previous = cumulative
        cumulative += cashflows[year]
        if cumulative >= 0 and cashflows[year] != 0:
            fraction = (0 - previous) / cashflows[year]
            payback = (year - 1) + max(0.0, min(1.0, fraction))
            break

    return npv, irr, payback, cashflows


def economic_analysis(fields, details, assumptions):
    rows, raw = [], {}
    for field in fields:
        r = details[field]
        npv, irr, payback, cashflows = project_cashflows(
            r["Actual CAPEX"], r["Actual OPEX"],
            assumptions["carbon_credit"], assumptions["government_subsidy"],
            assumptions["tax_incentive"], assumptions["storage_fee"],
            assumptions["discount_rate"], assumptions["inflation_rate"],
            assumptions["project_lifetime"], assumptions["injection_rate_mtpa"],
            assumptions["carbon_price"], assumptions.get("inflate_revenues", True),
        )
        raw[field] = {"NPV": npv, "IRR": irr, "Payback Period": payback, "Cashflows": cashflows}
        rows.append({
            "Field": field,
            "Actual CAPEX": r["Actual CAPEX"],
            "Actual OPEX": r["Actual OPEX"],
            "NPV": npv,
            "IRR (%)": None if irr is None else irr * 100,
            "Payback Period (years)": payback,
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df["_npv_rank"] = df["NPV"].rank(ascending=False, method="min")
        df["_irr_rank"] = df["IRR (%)"].rank(ascending=False, method="min", na_option="bottom")
        df["_payback_rank"] = df["Payback Period (years)"].rank(ascending=True, method="min", na_option="bottom")
        df["Economic Rank Score"] = (df["_npv_rank"] + df["_irr_rank"] + df["_payback_rank"]) / 3
        df = df.sort_values(["Economic Rank Score", "NPV", "IRR (%)"], ascending=[True, False, False], na_position="last").reset_index(drop=True)
        df.insert(0, "Rank", range(1, len(df) + 1))
        df = df.drop(columns=["_npv_rank", "_irr_rank", "_payback_rank"])
    return df, raw


def sensitivity_oat(fields, details, base_assumptions, ranges):
    rows, scenario_tables = [], {}
    for field in fields:
        scenarios = []
        for parameter, (low, high) in ranges.items():
            for label, value in [("Low", low), ("High", high)]:
                a = dict(base_assumptions)
                actual_capex = details[field]["Actual CAPEX"]
                actual_opex = details[field]["Actual OPEX"]

                if parameter == "CAPEX":
                    actual_capex = value
                elif parameter == "OPEX":
                    actual_opex = value
                elif parameter == "Discount Rate":
                    a["discount_rate"] = value
                elif parameter == "Inflation Rate":
                    a["inflation_rate"] = value
                elif parameter == "CO2 Injection Rate":
                    a["injection_rate_mtpa"] = value
                elif parameter == "Project Lifetime":
                    a["project_lifetime"] = int(round(value))
                elif parameter == "Carbon Credits":
                    a["carbon_credit"] = value

                npv, irr, payback, _ = project_cashflows(
                    actual_capex, actual_opex,
                    a["carbon_credit"], a["government_subsidy"], a["tax_incentive"],
                    a["storage_fee"], a["discount_rate"], a["inflation_rate"],
                    a["project_lifetime"], a["injection_rate_mtpa"],
                    a["carbon_price"], a.get("inflate_revenues", True),
                )
                scenarios.append({
                    "Field": field,
                    "Parameter": parameter,
                    "Case": label,
                    "Value": value,
                    "NPV": npv,
                    "IRR (%)": None if irr is None else irr * 100,
                    "Payback Period (years)": payback,
                })

        sdf = pd.DataFrame(scenarios)
        scenario_tables[field] = sdf
        irr_vals = sdf["IRR (%)"].dropna()
        pb_vals = sdf["Payback Period (years)"].dropna()
        rows.append({
            "Field": field,
            "Sensitivity NPV Min": sdf["NPV"].min(),
            "Sensitivity NPV Max": sdf["NPV"].max(),
            "Sensitivity IRR Min (%)": irr_vals.min() if not irr_vals.empty else None,
            "Sensitivity IRR Max (%)": irr_vals.max() if not irr_vals.empty else None,
            "Sensitivity Payback Min (years)": pb_vals.min() if not pb_vals.empty else None,
            "Sensitivity Payback Max (years)": pb_vals.max() if not pb_vals.empty else None,
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df["_npv_rank"] = df["Sensitivity NPV Min"].rank(ascending=False, method="min")
        df["_irr_rank"] = df["Sensitivity IRR Min (%)"].rank(ascending=False, method="min", na_option="bottom")
        df["_pb_rank"] = df["Sensitivity Payback Max (years)"].rank(ascending=True, method="min", na_option="bottom")
        df["Sensitivity Rank Score"] = (df["_npv_rank"] + df["_irr_rank"] + df["_pb_rank"]) / 3
        df = df.sort_values(["Sensitivity Rank Score", "Sensitivity NPV Min"], ascending=[True, False], na_position="last").reset_index(drop=True)
        df.insert(0, "Rank", range(1, len(df) + 1))
        df = df.drop(columns=["_npv_rank", "_irr_rank", "_pb_rank"])
    return df, scenario_tables
